import pandas as pd
import numpy as np
from openpyxl import load_workbook


import os

replacements = {
    1.0: 0.02,
    2.0: 0.03,
    3.0: 0.05,
    4.0: 0.07,
    5.0: 0.08,
    6.0: 0.10,
    7.0: 0.12,
    8.0: 0.13,
    9.0: 0.15,
    10.0: 0.17,
    11.0: 0.18,
    12.0: 0.20,
    13.0: 0.22,
    14.0: 0.23,
    15.0: 0.25,
    16.0: 0.27,
    17.0: 0.28,
    18.0: 0.30,
    19.0: 0.32,
    20.0: 0.33,
    21.0: 0.35,
    22.0: 0.37,
    23.0: 0.38,
    24.0: 0.40,
    25.0: 0.42,
    26.0: 0.43,
    27.0: 0.45,
    28.0: 0.47,
    29.0: 0.48,
    30.0: 0.50,
    31.0: 0.52,
    32.0: 0.53,
    33.0: 0.55,
    34.0: 0.57,
    35.0: 0.58,
    36.0: 0.60,
    37.0: 0.62,
    38.0: 0.63,
    39.0: 0.65,
    40.0: 0.67,
    41.0: 0.68,
    42.0: 0.70,
    43.0: 0.72,
    44.0: 0.73,
    45.0: 0.75,
    46.0: 0.77,
    47.0: 0.78,
    48.0: 0.80,
    49.0: 0.82,
    50.0: 0.83,
    51.0: 0.85,
    52.0: 0.87,
    53.0: 0.88,
    54.0: 0.90,
    55.0: 0.92,
    56.0: 0.93,
    57.0: 0.95,
    58.0: 0.97,
    59.0: 0.98,
    60.0: 1,
}

# Leer los datos del archivo de Excel




book = load_workbook(os.getcwd()+'/horas_trabajadas.xlsx')

#seleccionar la hoja de trabajo
sh = book.active


for row in sh.iter_rows():
    for cell in row:
        if cell.value:
            if isinstance(cell.value, (float, int)):
                decimal = round((cell.value*100)%100, 2)
                if decimal in replacements:
                    cell.value = int(cell.value) + replacements[decimal]

                                                                     


# Guardar los cambios en el mismo archivo excel
book.save(os.getcwd()+'/horas_trabajadas.xlsx')

input('Archivo Convertido')